def ten_coefficients(upload_folder='uploads', output_excel='10Coefficient.xlsx',units=None):
    import os
    import pandas as pd 
    import xlwings as xw
    from sklearn.model_selection import train_test_split
    from sklearn.linear_model import LinearRegression
    from openpyxl import load_workbook

        # ✅ Save output Excel to static folder
    static_folder = os.path.join(os.getcwd(), 'static')
    os.makedirs(static_folder, exist_ok=True)
    output_excel_path = os.path.join(static_folder, output_excel)


    def get_latest_file(folder, extension='.xlsm'):
        files = [f for f in os.listdir(folder) if f.endswith(extension)]
        if not files:
            raise FileNotFoundError(f"No '{extension}' files found in {folder}.")
        full_paths = [os.path.join(folder, f) for f in files]
        return max(full_paths, key=os.path.getmtime)

    def load_all_sheets(file_path):
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        all_sheets_data = {
            sheet.name.strip(): sheet.used_range.options(pd.DataFrame, header=False, index=False).value
            for sheet in wb.sheets
        }
        wb.close()
        app.quit()
        return all_sheets_data

    def reshape_to_long_format(df, name):
        evap_temps = df.iloc[4, 2:]
        cap = df.iloc[4:, 1:]
        cap_index = cap.iloc[1:, 0]
        cap_header = cap.iloc[0, 1:]
        cap = cap.iloc[1:, 1:]
        cap.columns = cap_header
        cap.index = cap_index
        cap_reset = cap.reset_index()
        melted = pd.melt(cap_reset, id_vars=cap_reset.columns[0], var_name='Te', value_name=name)
        melted.columns = ['Tc', 'Te', name]
        return melted[melted[name] != 0]

    def create_features(df, name):
        df["Te^2"] = df["Te"] ** 2
        df["Te*Tc"] = df["Te"] * df["Tc"]
        df["Tc^2"] = df["Tc"] ** 2
        df["Te^3"] = df["Te"] ** 3
        df["Te^2*Tc"] = (df["Te"] ** 2) * df["Tc"]
        df["Te*Tc^2"] = (df["Tc"] ** 2) * df["Te"]
        df["Tc^3"] = df["Tc"] ** 3
        return df[[name, 'Te', 'Tc', 'Te^2', 'Te*Tc', 'Tc^2', 'Te^3', 'Te^2*Tc', 'Te*Tc^2', 'Tc^3']]

    def train_and_get_coefficients(X, Y):
        X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.1, random_state=0)
        model = LinearRegression()
        model.fit(X_train, y_train)
        return [model.intercept_] + model.coef_.tolist()

    # Column mappings in output Excel
    column_map = {
        'Mass Flow Rate Table': 'E',
        'Power Table': 'F',
        'Capacity Table': 'I'
    }

    # Step 1: Get latest Excel file
    latest_file = get_latest_file(upload_folder)

    # Step 2: Load input Excel sheets
    all_sheets_data = load_all_sheets(latest_file)

    # Step 3: Prepare Excel output file
    if not os.path.exists(output_excel_path):
        from openpyxl import Workbook
        wb_new = Workbook()
        wb_new.save(output_excel_path)

    wb = load_workbook(output_excel_path)
    output_sheet_name = wb.sheetnames[0]
    sheet = wb[output_sheet_name]

    # Step 4: Loop through and write coefficients
    final_coefficients = {}

    for name, df in all_sheets_data.items():
        if name not in column_map:
            continue  # Skip sheets not in column map

        try:
            df_long = reshape_to_long_format(df, name)
            # 🟦 Unit Conversion for Temperature
            if units and units.get("temperature") == "degC":
                df_long["Te"] = df_long["Te"] * 9 / 5 + 32
                df_long["Tc"] = df_long["Tc"] * 9 / 5 + 32

            # 🟦 Unit Conversion for Output Value
            if units:
                if name == "Capacity Table" and units.get("capacity") == "W":
                    df_long[name] = df_long[name] * 3.412  # BTU/hr → W
                if name == "Mass Flow Rate Table" and units.get("massflow") == "kg/hr":
                    df_long[name] = df_long[name] * 2.20462  # lb/hr → kg/hr

            df_long = create_features(df_long, name)

            X, Y = df_long.iloc[:, 1:], df_long[name]
            coeffs = train_and_get_coefficients(X, Y)

            col_letter = column_map[name]
            for i, value in enumerate(coeffs):
                sheet[f'{col_letter}{5 + i}'] = value

            final_coefficients[name] = {f'{name[0].lower()}{i}': c for i, c in enumerate(coeffs)}
            print(df_long)
        except Exception as e:
            final_coefficients[name] = f"Error processing: {e}"

    wb.save(output_excel_path)
    # Actual coefficient terms
    term_labels = ["Intercept", "Te", "Tc", "Te^2", "Te*Tc", "Tc^2", "Te^3", "Te^2*Tc", "Te*Tc^2", "Tc^3"]

    # Fill dictionary from final_coefficients
    structured = {"Term": term_labels}
    for table in ["Capacity Table", "Power Table", "Mass Flow Rate Table"]:
        if table in final_coefficients:
            values = final_coefficients[table]
            if isinstance(values, dict):  # convert dict → list
                structured[table] = list(values.values())
            elif isinstance(values, list):
                structured[table] = values
            else:
                structured[table] = ["Error"] * 10
        else:
            structured[table] = ["Missing"] * 10

    return structured,output_excel_path
